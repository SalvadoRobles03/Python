from threading import Thread
import time
import openpyxl
import pandas as pd
import gspread
from df2gspread import df2gspread as d2g
import gspread_dataframe as gd
from oauth2client.service_account import ServiceAccountCredentials


scope=['https://www.googleapis.com/auth/spreadsheets',
       'https://www.googleapis.com/auth/drive.file',
       'https://www.googleapis.com/auth/drive']

creds=ServiceAccountCredentials.from_json_keyfile_name("Interprete.json",scope)

client=gspread.authorize(creds)

miLista=client.open("Subir de Excel").worksheet("Hoja 1")
miLista.update_cell(2,2,"NO")




def Timer(elección):
    if elección == "2":
      for i in range(16):
          minutos=0
          time.sleep(60)
          minutos = minutos+1
          if minutos==15:
              miLista.update_cell(2,2,"SI")
              print("\nTIEMPO TERMINADO")
              
              break
    if elección == "3":
      for i in range(16):
          minutos=0
          time.sleep(60)
          minutos = minutos+1
          if minutos==15:
              miLista.update_cell(2,2,"SI")
              print("\nTIEMPO TERMINADO")
              break
        
    if elección == "1":
        for i in range(2):
            minutos=0
            time.sleep(20)
            minutos = minutos+1
            if minutos==1:
                miLista.update_cell(2,2,"SI")
                print("\nTIEMPO TERMINADO")
                break

def separador():
    print("--------------------------------------------------")

def EleccionValida(numero):
    if numero == "1" or numero =="2" or numero =="3":
        return numero
    else:
        separador()
        print("\nIngrese una opción valida")
        EleccionValida(input("Elij@ la opción con la que desea empezar: "))
        
def ponerValorNo():
    miLista=client.open("Subir de Excel").worksheet("Hoja 1")
    miLista.update_cell(2,2,"NO")       

def main(elección):
    if elección == "1":
        ponerValorNo()
        separador()
        Thread(target = Timer, args=(elección)).start()
        print("Bienvenid@ a la sección de Matemáticas")
        Terminado=client.open("Subir de Excel").worksheet("Hoja 1").cell(2, 2).value
        while Terminado== "NO":
            Terminado=client.open("Subir de Excel").worksheet("Hoja 1").cell(2, 2).value
            time.sleep(5)
            print("SI JALA MATE")
        print("FIN DEL TEST")
    
    if elección == "2":
        ponerValorNo()
        separador()
        Thread(target = Timer, args=(elección)).start()
        print("Bienvenid@ a la sección de Ciencias Naturales")
        Terminado=client.open("Subir de Excel").worksheet("Hoja 1").cell(2, 2).value
        if Terminado== "NO":
            Terminado=client.open("Subir de Excel").worksheet("Hoja 1").cell(2, 2).value
            time.sleep(5)
            print("SI JALA CIENCIAS")
        else:
            print("FIN DEL TEST")

    if elección == "3":
        ponerValorNo()
        separador()
        Thread(target = Timer, args=(elección)).start()
        print("Bienvenid@ a la sección de Español")
        Terminado=client.open("Subir de Excel").worksheet("Hoja 1").cell(2, 2).value
        if Terminado== "NO":
            Terminado=client.open("Subir de Excel").worksheet("Hoja 1").cell(2, 2).value
            time.sleep(5)
            print("SI JALA ESPAÑOL")
        else:
            print("FIN DEL TEST")



print("Bienvenid@ a el servicio de estudio para exámenes profesionales")
print("El test consta de 3 secciones:")
print(" 1.Matemáticas\n 2.Ciencias Naturales\n 3.Español")

elección=EleccionValida(input("Elija la opción con la que desea empezar: "))
Terminado=client.open("Subir de Excel").worksheet("Hoja 1").cell(2, 2).value

main(elección)





